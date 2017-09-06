import datetime
from openpyxl import load_workbook

from flange.flange import Body, Bolt, RingJoint, Flange, Scenario
from flange.material import Material
from flange.design_cond import DesignCond

from flange.asme8 import Asme_Div1_Api, Asme_Div1_Asme


def main():

    # Extract inputs and assign to lists
    materials = collect_items('material')
    bodys = collect_items('body')
    bolts = collect_items('bolt')
    ringjoints = collect_items('ringjoint')
    des_conds = collect_items('des_cond')
    flanges = collect_items('flange')
    scenarios = collect_items('scenario')

    # Initialise results list
    analyses = []

    # Run analyses for all scenarios
    for scenario in scenarios:

        flange = flanges[scenario.flange]
        body = bodys[flange.body]
        body_mat = materials[flange.body_mat]
        bolt = bolts[flange.bolt]
        bolt_mat = materials[flange.bolt_mat]
        ringjoint = ringjoints[flange.ringjoint]
        des_cond = des_conds[scenario.des_cond]
        bolt_des_fac = scenario.bolt_des_fac

        codes = {'ASME': Asme_Div1_Asme, 'API': Asme_Div1_Api}

        analysis = codes[scenario.kind](scenario, flange, body, body_mat, bolt,
                                        bolt_mat, ringjoint, des_cond, bolt_des_fac)
        analyses.append(analysis)

    post_process(analyses)


def collect_items(item):
    '''
    Function to extract inputs from xlsx input file
    '''
    options = {'material': Material,
               'body': Body,
               'bolt': Bolt,
               'ringjoint': RingJoint,
               'des_cond': DesignCond,
               'flange': Flange,
               'scenario': Scenario}

    wb = load_workbook(filename='flange/inputs.xlsx')
    ws = wb.get_sheet_by_name(name=item)

    items = []

    count = 0
    for row in ws.iter_rows():
        # Condition to skip headers
        if count > 0:
            props = []

            for cell in row:
                props.append(cell.value)

            component = options[item](props)
            items.append(component)

        count += 1

    return items


def post_process(analyses):
    output = open('output.txt', 'w')
    output.write('ASME VIII Flange Design\n')
    output.write('%s\n' % datetime.datetime.today())
    output.write(
        '=========================================================================\n')

    for analysis in analyses:
        output.write(
            '-------------------------------------------------------------------------\n')
        output.write('Scenario %s\n' % analysis.ID)
        output.write(
            '-------------------------------------------------------------------------\n')

        output.write('Summary:\n')
        output.write('\tFlange Type: %s\n' % analysis.kind)
        output.write('\tFlange ID: %s\n' % analysis.flange_ID)
        output.write('\tDesign Condition ID: %s\n' % analysis.des_cond_ID)
        output.write('\tBolt Design Factor: %s\n\n' % analysis.bolt_des_fac)

        cases = ['Operational', 'Gasket Seating']
        urs = ['S_H', 'S_R', 'S_T', 'S_add1', 'S_add2']

        for case in cases:
            output.write("%s: \n" % case)
            for ur in urs:
                output.write("\t%s =\t%s\n" % (ur, analysis.urs[case][ur]))
            output.write("\n")

        output.write("Bolt UR =\t%s\n" % analysis.bolt_ur)
        output.write("\n")
        output.write("Max UR governed by %s condition = %s\n\n\n" %
                     (analysis.govern, analysis.ur_max))

    output.close()


if __name__ == "__main__":
    main()
